import csv
import glob
import inspect
import itertools
import logging
import datetime
import os
import random
import string
import time

import pandas as pd
from os import remove
from os.path import exists
from pathlib import Path
from random import randint
from xlutils.copy import copy
import pytest
import openpyxl
import xlrd
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook


@pytest.mark.usefixtures("browser_setup")
class BaseClass:
    download_folder = "download/folder"

    # Generate random account numbers
    def generate_acc_no(self, length):
        acc_number = ''.join(["{}".format(randint(0, 9)) for num in range(0, length)])
        return acc_number

    # Generate random characters
    def generate_random_char(self, length):
        letters = string.ascii_lowercase
        gen_char =  ''.join(random.choice(letters) for i in range(length))
        final_char = "TEST_" +gen_char
        return final_char

    # Open the Manual HU file from the path and update the 10 digit random account number
    def update_accno_excel_HU(self, fileloc, accno):
        book = openpyxl.load_workbook(fileloc)
        sheet = book.active
        sheet.cell(row=2, column=1).value = accno
        book.save(fileloc)

    # Open the excel file from the path and update the 'n' digit random account number in specified rows
    def update_accono_excel_multiple_rows(self, fileloc, accno, rownum):
        book = openpyxl.load_workbook(fileloc)
        sheet = book.active
        for i in range(2, rownum):
            sheet.cell(row=i, column=1).value = accno
        book.save(fileloc)

    def update_accono_excel_multiple_rows_emailHI(self, fileloc, accno, rownum):
        book = openpyxl.load_workbook(fileloc)
        sheet = book.active
        for i in range(2, rownum):
            sheet.cell(row=i, column=2).value = accno
        book.save(fileloc)

    # Handling .xls file and update the account number
    def update_accno_excel_xls(self, fileloc, accno, rownum):
        book = xlrd.open_workbook(fileloc)
        book1 = copy(book)
        sheet = book1.get_sheet(0)
        for i in range(1, rownum):
            sheet.write(i, 1, accno)
        book1.save(fileloc)

    # Handling .xls file and update the account number for Webscrape HU file
    def update_accno_excel_webscrape_HU(self, fileloc, accno):
        book = xlrd.open_workbook(fileloc)
        book1 = copy(book)
        sheet = book1.get_sheet(0)
        sheet.write(3, 0, "Column name: " + accno)
        book1.save(fileloc)

    # Handling csv file and update the account number
    def update_accno_csv_webscrape_HI(self, fileloc, accno):
        df = pd.read_csv(fileloc)  # Read the csv file
        pre_accno = df.loc[1, 'ACCOUNT_NO']  # Get the previous account number present in the file
        df['ACCOUNT_NO'] = df['ACCOUNT_NO'].replace({pre_accno: accno})  # Replace with newly generated
        df.to_csv(fileloc, index=False)  # Save the file

    # Generate multiple random account numbers
    def random_acc(self, n):
        def random_gen(low, high):
            while True:
                yield random.randrange(low, high)
        gen = random_gen(1000000000, 9999999999)
        items = list(itertools.islice(gen, n))  # Take first 10 random numbers
        return items

    # Update multiple account numbers for mass upload
    def update_multiple_accno_mass_update(self, fileloc, acc_numbers):
        book = openpyxl.load_workbook(fileloc)
        sheet = book.active
        for i in range(0, 10):
            # sheet.write(row = i+2, column = 1, items[i])
            sheet.cell(row=i + 2, column=1).value = acc_numbers[i]
        book.save(fileloc)

    # Logging function
    def getLogger(self):
        now = datetime.datetime.now()
        file_name = now.strftime('TestLog_%m%d%Y--%H-%M-%S.log')
        log_file = "results\\LogFiles\\"+file_name
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler(log_file)
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        fileHandler.setFormatter(formatter)
        logger.addHandler(fileHandler)
        logger.setLevel(logging.INFO)
        return logger

    # Function to get the Current URL
    def get_current_url(self):
        self.driver.current_url

    # Select drop down by it's value
    def select_by_value(self, locator, text):
        sel = Select(locator)
        sel.select_by_value(text)

    # Select drop down by it's index
    def select_by_index(self, locator, index):
        sel = Select(locator)
        sel.select_by_index(index)

    # Select drop down by it's visible text
    def select_by_text(self, locator, text):
        sel = Select(locator)
        sel.select_by_visible_text(text)

    # Wait for element to be present
    def wait_for_element_to_present(self, locator):
        WebDriverWait(self.driver, 60).until(EC.presence_of_element_located(locator))

    # Wait for element to be clickable
    def wait_for_element_to_clickable(self, locator):
        WebDriverWait(self.driver, 60).until(EC.element_to_be_clickable(locator))

    # Wait for element to be visible
    def wait_for_element_to_visible(self, locator):
        WebDriverWait(self.driver, 120).until(EC.visibility_of_element_located(locator))

    # Wait for all the elements to be visible
    def wait_for_all_elements_to_visible(self, locator):
        WebDriverWait(self.driver, 60).until(EC.visibility_of_all_elements_located(locator))

    # Wait for text in the element to be present
    def wait_for_text_in_ele_present(self, locator, text):
        WebDriverWait(self.driver, 60).until(EC.text_to_be_present_in_element(locator, text))

    # Wait for the element to be selected
    def wait_for_element_to_selected(self, locator):
        WebDriverWait(self.driver, 60).until(EC.element_to_be_selected(locator))

    # Fluent wait
    def wait_for_some_period(self, locator):
        element = None
        i = 6
        while element is None:
            try:
                wait = WebDriverWait(self.driver, 5, poll_frequency=1)
                element = wait.until(EC.visibility_of_element_located(locator))
            except:
                i = i - 1
                print(i)
                if i < 0:
                    raise Exception('Element not found')

    # Mouse hover function
    def move_to(self, locator):
        action = ActionChains(self.driver)
        action.move_to_element(locator).perform()

    # Mouse hover to the elements and click on the element
    def move_to_element_click(self, locator):
        action = ActionChains(self.driver)
        action.move_to_element(locator).click().perform()

    # Get text of the element
    def get_text_element(self, locator):
        ele = self.driver.find_element(locator)
        ele.text

    # Get the attribute value
    def get_attribute(self, locator, attribute_name):
        ele = self.driver.find_element(locator)
        ele.get_attribute(attribute_name)

    # Click on the element
    def click_element(self, locator):
        self.driver.find_element(locator).click()

    # Delete all the files in the Downloads folder
    def clear_downloads_folder(self):
        os.chdir(BaseClass.download_folder)
        files = glob.glob('*.*')
        for f in files:
            os.unlink(f)

    # Check the file is downloaded in system
    def is_download_finished(self):
        firefox_temp_file = sorted(Path(BaseClass.download_folder).glob('*.part'))
        chrome_temp_file = sorted(Path(BaseClass.download_folder).glob('*.crdownload'))
        downloaded_files = sorted(Path(BaseClass.download_folder).glob('*.*'))
        if (len(firefox_temp_file) == 0) and \
                (len(chrome_temp_file) == 0) and \
                (len(downloaded_files) >= 1):
            return True
        else:
            return False